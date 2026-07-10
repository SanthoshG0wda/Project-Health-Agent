from langchain.tools import tool
from pathlib import Path
import openpyxl
import pandas as pd
import datetime as dt


JUNK_VALUES = {"#UNPARSEABLE", "#REF!", "#N/A", "#VALUE!", "", None}


def _clean(value):
    """Return None if value is a known junk/error marker, else the value."""
    if isinstance(value, str) and value.strip() in JUNK_VALUES:
        return None
    if value in JUNK_VALUES:
        return None
    return value


def _find_task_sheet(workbook):
    """The task-level sheet name varies by export; pick the largest sheet
    that isn't 'Summary' or 'Comments'."""
    candidates = [
        ws for ws in workbook.worksheets
        if ws.title not in ("Summary", "Comments") and ws.max_row > 1
    ]
    if not candidates:
        return None
    return max(candidates, key=lambda ws: ws.max_row)

@tool
def read_excel_file(filepath: str):
    """Read excel file at the given path and return the contents as a pandas DataFrame."""
    if not Path(filepath).exists():
        return f"File not found: {filepath}"
    wb = pd.read_excel(filepath)
    return wb

@tool
def extract_project_data(filepath: str, as_of_date=None):
    """
    Returns a dict:
    {
        "project_name": str,
        "project_manager": str | None,
        "pct_complete": float | None,
        "schedule_health_summary": "Red"/"Amber"/"Green" | None,
        "at_risk_flag": str | None,
        "task_totals": {"not_started": int, "in_progress": int, "completed": int, "total": int},
        "task_schedule_health": {"Red": int, "Amber": int, "Green": int, "unlabeled": int},
        "overdue_incomplete_tasks": int,
        "overdue_critical_tasks": int,
        "budget": {"total": float|None, "spent": float|None},  # not present in this sample
        "stakeholder_sentiment": None,  # not present in this sample
        "missing_data": [list of field names that could not be extracted],
        "extraction_warnings": [list of human-readable notes],
    }
    """
    as_of_date = as_of_date or dt.datetime.now()
    missing = []
    warnings = []

    wb = openpyxl.load_workbook(filepath, data_only=True)

    # ---- Summary sheet (project-level metadata) ----
    project_name = None
    project_manager = None
    pct_complete = None
    schedule_health_summary = None
    at_risk_flag = None

    if "Summary" in wb.sheetnames:
        ws = wb["Summary"]
        summary = {}
        for row in ws.iter_rows(values_only=True):
            if row and row[0] is not None:
                summary[row[0]] = _clean(row[1]) if len(row) > 1 else None

        project_name = summary.get("Project Name")
        project_manager = summary.get("Project Manager")
        pct_complete = summary.get("% Complete")
        schedule_health_summary = summary.get("Schedule Health")
        at_risk_flag = summary.get("At Risk")
    else:
        warnings.append("No 'Summary' sheet found; project-level metadata unavailable.")
        missing.append("summary_metadata")

    # ---- Task sheet (task-level rollups) ----
    task_ws = _find_task_sheet(wb)
    task_totals = {"not_started": 0, "in_progress": 0, "completed": 0, "total": 0}
    task_schedule_health = {"Red": 0, "Amber": 0, "Green": 0, "unlabeled": 0}
    overdue_incomplete = 0
    overdue_critical = 0
    is_simple_format = False

    if task_ws is None:
        warnings.append("No task-level sheet found; task rollups unavailable.")
        missing.append("task_level_data")
    else:
        headers = [c.value for c in next(task_ws.iter_rows(min_row=1, max_row=1))]
        idx = {h: i for i, h in enumerate(headers) if h is not None}

        required_cols = ["Status", "Schedule Health", "End Date", "Task Name"]
        missing_cols = [c for c in required_cols if c not in idx]

        # Check if this is a simple flat format instead
        if missing_cols and "Project" in idx and "Budget" in idx:
            is_simple_format = True
        else:
            if missing_cols:
                warnings.append(f"Task sheet missing expected columns: {missing_cols}")

        first_task_name = None
        for i, row in enumerate(task_ws.iter_rows(min_row=2, values_only=True)):
            def get(col):
                return _clean(row[idx[col]]) if col in idx and idx[col] < len(row) else None

            task_totals["total"] += 1

            if is_simple_format:
                continue

            status = get("Status")
            if status == "Not Started":
                task_totals["not_started"] += 1
            elif status == "In Progress":
                task_totals["in_progress"] += 1
            elif status == "Completed":
                task_totals["completed"] += 1

            # Normalize schedule health naming (source uses "Yellow", we use "Amber")
            sh = get("Schedule Health")
            if sh == "Yellow":
                sh = "Amber"
            if sh in ("Red", "Amber", "Green"):
                task_schedule_health[sh] += 1
            else:
                task_schedule_health["unlabeled"] += 1

            end_date = get("End Date")
            critical = get("Critical ?") if "Critical ?" in idx else None
            if isinstance(end_date, dt.datetime) and end_date < as_of_date and status != "Completed":
                overdue_incomplete += 1
                if critical:
                    overdue_critical += 1

            if first_task_name is None:
                tn = get("Task Name")
                if tn:
                    first_task_name = tn

        if not project_name and first_task_name:
            project_name = first_task_name
            warnings.append("Project Name missing from Summary sheet; used first task row's name as fallback.")

    # ---- Simple flat format (Project, Start Date, End Date, Completion, Budget, Spent, ...) ----
    budget = {"total": None, "spent": None}
    stakeholder_sentiment = None

    if is_simple_format:
        ws = wb.active
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        idx = {h: i for i, h in enumerate(headers) if h is not None}
        if "Project" in idx and "Budget" in idx:
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            for ri, row in enumerate(rows):
                if ri == 0:
                    def get(col):
                        return _clean(row[idx[col]]) if col in idx and idx[col] < len(row) else None
                    pn = get("Project")
                    if pn:
                        project_name = pn
                    pct_raw = get("Completion")
                    if pct_raw:
                        try:
                            pct_complete = float(str(pct_raw).replace("%", "")) / 100
                        except (ValueError, TypeError):
                            pass
                    budget_raw = get("Budget")
                    if budget_raw:
                        try:
                            budget["total"] = float(str(budget_raw).replace("$", "").replace(",", ""))
                        except (ValueError, TypeError):
                            pass
                    spent_raw = get("Spent")
                    if spent_raw:
                        try:
                            budget["spent"] = float(str(spent_raw).replace("$", "").replace(",", ""))
                        except (ValueError, TypeError):
                            pass
                    sentiment_raw = get("Manager Notes")
                    if sentiment_raw and sentiment_raw != "-":
                        stakeholder_sentiment = sentiment_raw
                    warnings.append("Detected simple flat format; extracting available fields.")
                else:
                    # Count additional rows as extra projects found
                    warnings.append(f"Extra project row {ri+1} found in simple format (only first project extracted).")

    if not project_name:
        project_name = "UNKNOWN PROJECT"
        missing.append("project_name")

    if pct_complete is None:
        missing.append("pct_complete")
    if schedule_health_summary is None:
        missing.append("schedule_health_summary")

    if budget["total"] is None:
        missing.append("budget")
    if stakeholder_sentiment is None:
        missing.append("stakeholder_sentiment")

    return {
        "project_name": project_name,
        "project_manager": project_manager,
        "pct_complete": pct_complete,
        "schedule_health_summary": schedule_health_summary,
        "at_risk_flag": at_risk_flag,
        "task_totals": task_totals,
        "task_schedule_health": task_schedule_health,
        "overdue_incomplete_tasks": overdue_incomplete,
        "overdue_critical_tasks": overdue_critical,
        "budget": budget,
        "stakeholder_sentiment": stakeholder_sentiment,
        "missing_data": missing,
        "extraction_warnings": warnings,
        "as_of_date": as_of_date.isoformat(),
    }